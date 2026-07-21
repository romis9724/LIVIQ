"""세대 마스터 시드 — xlsx(households 시트)로 buildings·households 일괄 등록 (H7-7).

명부 업로드(FR-ONB-07)는 세대(동·층·호)가 DB에 있어야 반영된다. 세대 마스터 관리 UI는
Phase 2(docs/04 단지 관리)라, 파일럿 단지 개설 시 이 스크립트로 세대를 시드한다.

입력 xlsx의 `households` 시트: 헤더 `household_id|dong|ho|floor|...`(1행) —
dong "401동"·ho "201호" 접미사는 제거해 building name "401"·unit_no 201로 저장한다.
멱등: 같은 (동, 층, 호) 세대가 이미 있으면 건너뛴다.

실행법(루트 .env 로드):

    cd apps/api
    uv run --no-sync --env-file .env python scripts/seed_households_xlsx.py \\
        --tenant-id 11111111-1111-1111-1111-111111111111 \\
        --file /path/to/단지_세대.xlsx
"""

from __future__ import annotations

import argparse
import asyncio
import uuid
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from liviq_db.engine import create_engine, create_session_factory
from liviq_db.models import Building, Household, Tenant

SHEET_NAME = "households"


def _parse_households(path: Path) -> dict[str, list[tuple[int, int]]]:
    """동 이름 → [(층, 호)] 목록. dong/ho의 '동'/'호' 접미사·공백은 제거."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"'{SHEET_NAME}' 시트가 없습니다 — 시트: {workbook.sheetnames}")
    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(min_row=2, values_only=True)

    by_dong: dict[str, list[tuple[int, int]]] = defaultdict(list)
    for cells in rows:
        if cells is None or all(c is None for c in cells[:4]):
            continue
        _, dong, ho, floor = cells[:4]
        if dong is None or ho is None or floor is None:
            continue
        name = str(dong).replace("동", "").strip()
        unit_no = int(str(ho).replace("호", "").strip())
        by_dong[name].append((int(floor), unit_no))
    workbook.close()
    if not by_dong:
        raise SystemExit("세대 행이 없습니다 — households 시트를 확인하세요")
    return dict(by_dong)


async def _seed(
    session: AsyncSession, tenant_id: uuid.UUID, by_dong: dict[str, list[tuple[int, int]]]
) -> tuple[int, int, int]:
    """buildings·households 멱등 upsert. (동 신규, 세대 신규, 세대 기존) 카운트 반환."""
    if await session.scalar(select(Tenant.id).where(Tenant.id == tenant_id)) is None:
        raise SystemExit(f"단지를 찾을 수 없습니다: {tenant_id}")
    await session.execute(
        text("SELECT set_config('app.tenant_id', :t, true)").bindparams(t=str(tenant_id))
    )

    buildings_created = created = skipped = 0
    for name, units in sorted(by_dong.items()):
        building_id = await session.scalar(
            select(Building.id).where(Building.tenant_id == tenant_id, Building.name == name)
        )
        if building_id is None:
            building_id = uuid.uuid4()
            session.add(
                Building(
                    id=building_id,
                    tenant_id=tenant_id,
                    name=name,
                    floors=max(floor for floor, _ in units),
                )
            )
            await session.flush()
            buildings_created += 1

        existing = {
            (floor, unit_no)
            for floor, unit_no in await session.execute(
                select(Household.floor, Household.unit_no).where(
                    Household.tenant_id == tenant_id, Household.building_id == building_id
                )
            )
        }
        for floor, unit_no in units:
            if (floor, unit_no) in existing:
                skipped += 1
                continue
            session.add(
                Household(
                    tenant_id=tenant_id,
                    building_id=building_id,
                    floor=floor,
                    unit_no=unit_no,
                    status="active",
                )
            )
            created += 1
        await session.flush()
    return buildings_created, created, skipped


async def _run(tenant_id: uuid.UUID, path: Path) -> None:
    by_dong = _parse_households(path)
    engine = create_engine()
    factory = create_session_factory(engine)
    try:
        async with factory() as session, session.begin():
            buildings, created, skipped = await _seed(session, tenant_id, by_dong)
        total = sum(len(u) for u in by_dong.values())
        print(
            f"동 {len(by_dong)}개(신규 {buildings}) · 세대 {total}건 — "
            f"신규 {created}, 기존 {skipped}"
        )
    finally:
        await engine.dispose()


def main() -> None:
    parser = argparse.ArgumentParser(description="세대 마스터 xlsx 시드(H7-7)")
    parser.add_argument("--tenant-id", required=True, type=uuid.UUID, help="대상 단지 UUID")
    parser.add_argument("--file", required=True, type=Path, help="households 시트를 담은 xlsx")
    args = parser.parse_args()
    if not args.file.exists():
        raise SystemExit(f"파일이 없습니다: {args.file}")
    asyncio.run(_run(args.tenant_id, args.file))


if __name__ == "__main__":
    main()

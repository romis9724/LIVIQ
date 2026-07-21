"""roster — 명부 엑셀 업로드 + 사전등록 diff 병합 (docs/01 §13, docs/03 §4.1, docs/11 §3.4).

MANAGER 전용. 컬럼 성함|생년월일|동|층|호(헤더 1행). 행 검증 실패는 error_report에 축적하고
나머지는 적용한다. diff 병합 키=(name_hash, birth_date_hash, household_id):
신규는 pre_registered 생성, 명부에서 사라진 pre_registered는 inactive 표시(자동 삭제 금지),
이미 가입(active 등)한 세대는 불변.
"""

from __future__ import annotations

import datetime
import io
import uuid
from typing import Annotated

from fastapi import APIRouter, Depends, File, HTTPException, Response, UploadFile
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field, ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.deps import RequestContext, Storage, get_storage, get_tenant_session, require_roles
from app.pii import PiiCrypto, get_pii_crypto
from app.routers.approvals import mask_name
from app.schemas.roster import (
    RosterCounts,
    RosterEntry,
    RosterLastUpload,
    RosterListOut,
    RosterRowError,
    RosterStateIn,
    RosterUploadOut,
)
from liviq_db.models import Building, ExcelUpload, Household, PiiVault, User

router = APIRouter(prefix="/admin/roster", tags=["roster"])

EXPECTED_HEADER = ("성함", "생년월일", "동", "층", "호")
MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 명부 엑셀 크기 상한


# 양식 다운로드 예시 행 — 실제 파서(EXPECTED_HEADER·RosterRowIn)와 형식이 항상 일치해야
# 한다(라운드트립 테스트로 보증, H7-7). 동은 문자열, 층·호는 정수.
_TEMPLATE_EXAMPLE_ROWS = (
    ("홍길동", "1980-01-15", "101", 3, 301),
    ("김영희", "1955-07-02", "101", 3, 302),
    ("이철수", "1992-11-30", "102", 5, 501),
)


@router.get("/template")
async def roster_template(
    _ctx: Annotated[RequestContext, Depends(require_roles("MANAGER"))],
) -> Response:
    """명부 업로드 양식 xlsx — 헤더+예시 행(H7-7). 다운로드 후 예시를 지우고 채워 쓴다."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None  # openpyxl 신규 워크북은 항상 활성 시트 보유
    sheet.title = "명부"
    sheet.append(list(EXPECTED_HEADER))
    for row in _TEMPLATE_EXAMPLE_ROWS:
        sheet.append(list(row))
    buffer = io.BytesIO()
    workbook.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="liviq-roster-template.xlsx"'},
    )


# 명부 행 상태(H7-9) — 판정은 전부 데이터에서 파생(별도 컬럼 없음):
# 미가입 = pre_registered·비삭제 · 가입완료 = pre_registered·소진(soft delete, 온보딩 매칭)
# 전출 후보 = inactive·비삭제(재업로드 diff에서 사라진 행)
_STATE_UNREGISTERED = "unregistered"
_STATE_JOINED = "joined"
_STATE_MOVED_OUT = "moved_out"
MAX_PAGE_SIZE = 200


def _entry_state(status: str, deleted: bool) -> str:
    if status == "inactive":
        return _STATE_MOVED_OUT
    return _STATE_JOINED if deleted else _STATE_UNREGISTERED


@router.get("", response_model=RosterListOut)
async def list_roster(
    ctx: Annotated[RequestContext, Depends(require_roles("MANAGER"))],
    session: Annotated[AsyncSession, Depends(get_tenant_session)],
    crypto: Annotated[PiiCrypto, Depends(get_pii_crypto)],
    q: str = "",
    state: str = "",
    page: int = 1,
    size: int = 50,
) -> RosterListOut:
    """명부 목록(H7-9) — 동·호·성함(마스킹)·상태 + 총계 + 마지막 업로드 요약.

    명부 행 = 명부 출신 사용자(login_id 없음·PII 참조 보유·pre_registered/inactive).
    q는 동 이름 또는 호수 일치 검색. 생년월일은 반환하지 않는다(운영자 결정, H7-9).
    """
    page = max(page, 1)
    size = min(max(size, 1), MAX_PAGE_SIZE)

    roster_rows = (
        select(
            User.id,
            User.status,
            User.deleted_at,
            PiiVault.name_enc,
            Building.name.label("building_name"),
            Household.floor,
            Household.unit_no,
        )
        .join(PiiVault, PiiVault.id == User.pii_ref)
        .outerjoin(Household, Household.id == User.household_id)
        .outerjoin(Building, Building.id == Household.building_id)
        .where(
            User.tenant_id == ctx.tenant_id,
            User.login_id.is_(None),
            User.pii_ref.is_not(None),
            User.status.in_(("pre_registered", "inactive")),
        )
    )

    # 총계는 필터와 무관하게 전체 기준(화면 상단 배지).
    all_rows = (await session.execute(roster_rows)).all()
    counts = RosterCounts(total=len(all_rows), unregistered=0, joined=0, moved_out=0)
    for row in all_rows:
        entry_state = _entry_state(row.status, row.deleted_at is not None)
        if entry_state == _STATE_UNREGISTERED:
            counts.unregistered += 1
        elif entry_state == _STATE_JOINED:
            counts.joined += 1
        else:
            counts.moved_out += 1

    query = q.strip()
    filtered = [
        row
        for row in all_rows
        if (not state or _entry_state(row.status, row.deleted_at is not None) == state)
        and (
            not query
            or row.building_name == query
            or (row.unit_no is not None and str(row.unit_no) == query)
        )
    ]
    filtered.sort(
        key=lambda r: (r.building_name or "", r.floor or 0, r.unit_no or 0, r.status)
    )
    page_rows = filtered[(page - 1) * size : page * size]

    dek = await crypto.get_dek(session, ctx.tenant_id) if page_rows else b""

    def masked(name_enc: bytes | None) -> str:
        if name_enc is None:
            return "*"
        try:
            return mask_name(crypto.decrypt(dek, name_enc))
        except Exception:  # noqa: BLE001 — 복호 실패 행도 목록엔 남긴다
            return "*"

    last = (
        await session.execute(
            select(ExcelUpload.created_at, ExcelUpload.row_count, ExcelUpload.error_report)
            .where(ExcelUpload.tenant_id == ctx.tenant_id, ExcelUpload.type == "roster")
            .order_by(ExcelUpload.created_at.desc())
            .limit(1)
        )
    ).first()

    return RosterListOut(
        items=[
            RosterEntry(
                user_id=row.id,
                name_masked=masked(row.name_enc),
                building_name=row.building_name,
                floor=row.floor,
                unit_no=row.unit_no,
                state=_entry_state(row.status, row.deleted_at is not None),
            )
            for row in page_rows
        ],
        total=len(filtered),
        counts=counts,
        last_upload=(
            RosterLastUpload(
                uploaded_at=last.created_at,
                row_count=last.row_count or 0,
                error_count=len((last.error_report or {}).get("errors", [])),
            )
            if last is not None
            else None
        ),
    )


async def _get_roster_row(
    session: AsyncSession, tenant_id: uuid.UUID, user_id: uuid.UUID
) -> User:
    """비삭제 명부 행(가입 계정 아님) — 없음·가입완료(소진)·타 단지는 동일 404."""
    user = await session.scalar(
        select(User).where(
            User.tenant_id == tenant_id,
            User.id == user_id,
            User.login_id.is_(None),
            User.pii_ref.is_not(None),
            User.status.in_(("pre_registered", "inactive")),
            User.deleted_at.is_(None),
        )
    )
    if user is None:
        raise HTTPException(status_code=404, detail="명부 행을 찾을 수 없습니다")
    return user


@router.patch("/{user_id}", status_code=204)
async def update_roster_state(
    user_id: uuid.UUID,
    body: RosterStateIn,
    ctx: Annotated[RequestContext, Depends(require_roles("MANAGER"))],
    session: Annotated[AsyncSession, Depends(get_tenant_session)],
) -> Response:
    """명부 행 상태 수동 변경(H7-9 보강) — 미가입 ↔ 전출 후보(소장 판단)."""
    if body.state not in (_STATE_UNREGISTERED, _STATE_MOVED_OUT):
        raise HTTPException(status_code=422, detail="상태는 미가입·전출 후보만 가능합니다")
    user = await _get_roster_row(session, ctx.tenant_id, user_id)
    user.status = "pre_registered" if body.state == _STATE_UNREGISTERED else "inactive"
    await session.flush()
    return Response(status_code=204)


@router.delete("/{user_id}", status_code=204)
async def delete_roster_row(
    user_id: uuid.UUID,
    ctx: Annotated[RequestContext, Depends(require_roles("MANAGER"))],
    session: Annotated[AsyncSession, Depends(get_tenant_session)],
) -> Response:
    """명부 행 삭제(H7-9 보강) — 가입 계정이 아닌 사전등록 행이라 PII vault째 완전 삭제."""
    user = await _get_roster_row(session, ctx.tenant_id, user_id)
    vault_id = user.pii_ref
    await session.delete(user)
    await session.flush()
    if vault_id is not None:
        vault = await session.scalar(
            select(PiiVault).where(PiiVault.id == vault_id, PiiVault.tenant_id == ctx.tenant_id)
        )
        if vault is not None:
            await session.delete(vault)
            await session.flush()
    return Response(status_code=204)


class RosterRowIn(BaseModel):
    """엑셀 한 행 — 경계 검증(docs 규칙 4). 동은 문자열(예: '101'), 층·호는 정수."""

    name: str = Field(min_length=1, max_length=64)
    birth_date: datetime.date
    building_name: str = Field(min_length=1, max_length=64)
    floor: int
    unit_no: int


@router.post("/upload", response_model=RosterUploadOut)
async def upload_roster(
    ctx: Annotated[RequestContext, Depends(require_roles("MANAGER"))],
    session: Annotated[AsyncSession, Depends(get_tenant_session)],
    storage: Annotated[Storage, Depends(get_storage)],
    crypto: Annotated[PiiCrypto, Depends(get_pii_crypto)],
    file: Annotated[UploadFile, File()],
) -> RosterUploadOut:
    data = await file.read()
    if not data:
        raise HTTPException(status_code=422, detail="빈 파일")
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="파일이 10MB를 초과")

    parsed, errors = _parse_rows(data)

    dek = await crypto.get_dek(session, ctx.tenant_id)
    existing = await _load_pre_registered(session, ctx.tenant_id)
    seen: set[tuple[str, str, uuid.UUID]] = set()
    applied = 0
    for row_no, row in parsed:
        household_id = await _lookup_household(session, ctx.tenant_id, row)
        if household_id is None:
            errors.append(RosterRowError(row=row_no, reason="해당 세대 없음"))
            continue
        key = (
            crypto.hmac_hash(row.name),
            crypto.hmac_hash(row.birth_date.isoformat()),
            household_id,
        )
        seen.add(key)
        if key in existing:
            continue  # 이미 사전등록됨 — 불변
        _create_pre_registered(session, crypto, dek, ctx.tenant_id, household_id, row, key)
        applied += 1

    marked_inactive = _mark_stale_inactive(existing, seen)

    upload_id = uuid.uuid4()
    file_key = f"{ctx.tenant_id}/roster/{upload_id}.xlsx"
    await storage.put(file_key, data)
    session.add(
        ExcelUpload(
            id=upload_id,
            tenant_id=ctx.tenant_id,
            type="roster",
            file_key=file_key,
            status="applied",
            row_count=len(parsed),
            error_report={"errors": [e.model_dump() for e in errors]} if errors else None,
            uploaded_by=ctx.user_id,
        )
    )
    await session.flush()
    return RosterUploadOut(
        upload_id=upload_id, applied=applied, marked_inactive=marked_inactive, errors=errors
    )


def _parse_rows(data: bytes) -> tuple[list[tuple[int, RosterRowIn]], list[RosterRowError]]:
    """xlsx 첫 시트 파싱. 헤더 검증 후 데이터 행을 Pydantic으로 검증."""
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl은 다양한 예외 — 손상 파일은 422
        raise HTTPException(status_code=422, detail="엑셀 파일을 읽을 수 없습니다") from exc
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if (
        header is None
        or tuple(str(c).strip() if c is not None else "" for c in header[:5]) != EXPECTED_HEADER
    ):
        workbook.close()
        raise HTTPException(
            status_code=422, detail=f"헤더는 {'|'.join(EXPECTED_HEADER)} 여야 합니다"
        )

    parsed: list[tuple[int, RosterRowIn]] = []
    errors: list[RosterRowError] = []
    for offset, cells in enumerate(rows):
        row_no = offset + 2  # 헤더=1
        if cells is None or all(c is None for c in cells[:5]):
            continue  # 빈 행 건너뜀
        try:
            parsed.append((row_no, _row_from_cells(cells)))
        except ValidationError as exc:
            errors.append(RosterRowError(row=row_no, reason=_first_error(exc)))
    workbook.close()
    return parsed, errors


def _row_from_cells(cells: tuple[object, ...]) -> RosterRowIn:
    name, birth, building, floor, unit = cells[:5]
    return RosterRowIn(
        name=str(name).strip() if name is not None else "",
        birth_date=birth,  # type: ignore[arg-type]  # Pydantic이 date/datetime/문자열 강제
        building_name=str(building).strip() if building is not None else "",
        floor=floor,  # type: ignore[arg-type]
        unit_no=unit,  # type: ignore[arg-type]
    )


def _first_error(exc: ValidationError) -> str:
    err = exc.errors()[0]
    field = ".".join(str(p) for p in err["loc"]) or "행"
    return f"{field}: {err['msg']}"


async def _load_pre_registered(
    session: AsyncSession, tenant_id: uuid.UUID
) -> dict[tuple[str, str, uuid.UUID], User]:
    rows = await session.execute(
        select(User, PiiVault.name_hash, PiiVault.birth_date_hash)
        .join(PiiVault, PiiVault.id == User.pii_ref)
        .where(
            User.tenant_id == tenant_id,
            User.status == "pre_registered",
            User.deleted_at.is_(None),
        )
    )
    result: dict[tuple[str, str, uuid.UUID], User] = {}
    for user, name_hash, birth_hash in rows:
        if name_hash and birth_hash and user.household_id:
            result[(name_hash, birth_hash, user.household_id)] = user
    return result


async def _lookup_household(
    session: AsyncSession, tenant_id: uuid.UUID, row: RosterRowIn
) -> uuid.UUID | None:
    return await session.scalar(
        select(Household.id)
        .join(Building, Building.id == Household.building_id)
        .where(
            Household.tenant_id == tenant_id,
            Building.name == row.building_name,
            Household.floor == row.floor,
            Household.unit_no == row.unit_no,
        )
    )


def _create_pre_registered(
    session: AsyncSession,
    crypto: PiiCrypto,
    dek: bytes,
    tenant_id: uuid.UUID,
    household_id: uuid.UUID,
    row: RosterRowIn,
    key: tuple[str, str, uuid.UUID],
) -> None:
    name_hash, birth_hash, _ = key
    vault_id = uuid.uuid4()
    session.add(
        PiiVault(
            id=vault_id,
            tenant_id=tenant_id,
            name_enc=crypto.encrypt(dek, row.name),
            birth_date_enc=crypto.encrypt(dek, row.birth_date.isoformat()),
            name_hash=name_hash,
            birth_date_hash=birth_hash,
            key_version=1,
        )
    )
    session.add(
        User(
            tenant_id=tenant_id,
            household_id=household_id,
            login_id=None,
            status="pre_registered",
            roster_matched=False,
            pii_ref=vault_id,
        )
    )


def _mark_stale_inactive(
    existing: dict[tuple[str, str, uuid.UUID], User],
    seen: set[tuple[str, str, uuid.UUID]],
) -> int:
    """명부에서 사라진 pre_registered 행 = inactive(전출 추정, 자동 삭제 금지)."""
    marked = 0
    for key, user in existing.items():
        if key not in seen:
            user.status = "inactive"
            marked += 1
    return marked

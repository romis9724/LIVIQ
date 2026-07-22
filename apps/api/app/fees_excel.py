"""관리비 엑셀 파싱·분배 — 순수 로직(DB 없음, H8-7 단지 총액 트리 계약).

입력 = 단지 총액 트리(첫 시트 2열): A열=분류(들여쓰기 트리), B열=금액.
들여쓰기 2칸당 depth 1(level = leading_spaces // 2). 헤더행(A='분류')·빈 A행은 스킵.
금액은 KRW 원 단위 정수(음수 허용 — 공용정산). 세대 매칭·적재는 라우터(DB) 소관.

분배(divide_fee_tree)는 **코드가 계산**한다(AI 미개입 → 규칙 5 위반 아님). 각 행을
독립적으로 세대수로 나눈다(ROUND_HALF_UP) — 부모/자식 재계산 없음(반올림 드리프트 수원, 데모 허용).
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook

HEADER_LABEL = "분류"  # 헤더행 A열 — 데이터가 아니므로 스킵


@dataclass(frozen=True)
class FeeTreeRow:
    level: int  # 들여쓰기 depth(0=대분류)
    name: str  # 항목명(공백 제거)
    amount: int  # 단지 총액(원 단위 정수, 음수 허용)


def parse_fee_total_xlsx(data: bytes) -> list[FeeTreeRow]:
    """xlsx 첫 시트 = 단지 총액 트리. 손상·형식 오류는 422, 헤더·빈 행은 스킵."""
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl은 다양한 예외 — 손상 파일은 422
        raise HTTPException(status_code=422, detail="엑셀 파일을 읽을 수 없습니다") from exc
    try:
        sheet = workbook.worksheets[0]
        rows: list[FeeTreeRow] = []
        for offset, cells in enumerate(sheet.iter_rows(values_only=True)):
            row_no = offset + 1  # 엑셀 행 번호(1-기반)
            raw_name = cells[0] if cells else None
            if raw_name is None or not str(raw_name).strip():
                continue  # 빈 분류행 스킵
            name_cell = str(raw_name)
            if name_cell.strip() == HEADER_LABEL:
                continue  # 헤더행 스킵
            level = _leading_spaces(name_cell) // 2
            amount = _to_amount(cells[1] if len(cells) > 1 else None)
            if amount is None:
                raise HTTPException(
                    status_code=422,
                    detail=f"{row_no}행 '{name_cell.strip()}': 금액이 정수여야 합니다",
                )
            rows.append(FeeTreeRow(level=level, name=name_cell.strip(), amount=amount))
        if not rows:
            raise HTTPException(status_code=422, detail="분배할 관리비 항목이 없습니다")
        return rows
    finally:
        workbook.close()


def divide_fee_tree(rows: list[FeeTreeRow], household_count: int) -> list[dict[str, Any]]:
    """각 행을 household_count로 독립 분배(ROUND_HALF_UP, 정수 원). 순서 보존 리스트 반환."""
    if household_count <= 0:
        raise ValueError("household_count는 양수여야 합니다")
    divisor = Decimal(household_count)
    return [
        {
            "name": row.name,
            "level": row.level,
            "amount": int(
                (Decimal(row.amount) / divisor).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
            ),
        }
        for row in rows
    ]


def _leading_spaces(text: str) -> int:
    return len(text) - len(text.lstrip(" "))


def _to_amount(cell: object) -> int | None:
    """금액 = 원 단위 정수(음수 허용). bool·비숫자·소수는 거절."""
    if isinstance(cell, bool):  # 엑셀 TRUE/FALSE 오입력 방지
        return None
    if isinstance(cell, int):
        return cell
    if isinstance(cell, float) and cell.is_integer():
        return int(cell)
    return None
